
import openpyxl as xl
import json
import re
import pandas as pd

from dataclasses import dataclass, field
from preprocessor import handle_duplicates, calc_grades, handle_zero_credit, handle_zero_on_fail

import os
import sys

total_subjects = {'Computer Fundamentals',
'Introduction to Linear Algebra',
'English Language',
'Electric Circuits',
'Structured Programming',
'Mathematics-1',
'Probability and Statistics-1',
'Concepts in Artificial Intelligence',
'Introduction to Programming with Python',
'Probability & Statistics 2',
'Scientific Thinking',
'Societal Issues',
'Logic Design',
'Numerical Analysis',
'Introduction to Machine Learning',
'Introduction to Computer Vision and Robotics',
'Object Oriented Programming',
'Computer Architecture',
'Computer Networks',
'Scientific Writing',
'Information Retrieval and Web Search',
'Databases',
'Fundamentals of Computational Intelligence',	
'Introduction to Data Structures',
'Introduction to Natural Language Processing',
'Operating Systems',
'Introduction to Multi Agent Systems Design',
'Computer Security',
'Fundamentals of Computer Graphics'}


dependecy_map = {
    'Concepts in Artificial Intelligence' : 'Computer Fundamentals',
    'Introduction to Programming with Python' : 'Structured Programming',
    'Probability & Statistics 2' : 'Probability and Statistics-1',
    'Logic Design' : 'Electric Circuits',
    'Numerical Analysis' : 'Introduction to Linear Algebra',
    'Introduction to Machine Learning' : 'Concepts in Artificial Intelligence',
    'Introduction to Computer Vision and Robotics' : 'Concepts in Artificial Intelligence',
    'Object Oriented Programming' : 'Structured Programming',
    'Computer Architecture' : 'Logic Design',
    'Computer Networks' : 'Computer Fundamentals',
    'Scientific Writing' : 'English Language',
    'Information Retrieval and Web Search' : 'Structured Programming',
    'Databases' : 'Structured Programming',
    'Fundamentals of Computational Intelligence' : 'Introduction to Machine Learning',
    'Introduction to Data Structures' : 'Object Oriented Programming',
    'Introduction to Natural Language Processing' : 'Concepts in Artificial Intelligence',
    'Operating Systems' : 'Computer Architecture',
    'Introduction to Multi Agent Systems Design' : 'Concepts in Artificial Intelligence',
    'Computer Security' :  'Computer Networks',
    'Fundamentals of Computer Graphics' : 'Introduction to Computer Vision and Robotics'}

next_year_lockers = ['Introduction to Computer Vision and Robotics', 'Introduction to Machine Learning', 'Computer Architecture',
                     'Introduction to Data Structures', 'Object Oriented Programming', 'Introduction to Data Structures', 'Computer Networks', 'Object Oriented Programming',
                     'Fundamentals of Computational Intelligence']



from collections import defaultdict

def get_dependency(name):
        if name in dependecy_map.keys():
            return [name] + get_dependency(dependecy_map[name])
        else:
            return []

def get_dependent_subjects(name, openned):
        name = dependecy_map.get(name)
        if name == None:
            return[]
        
        dependencies = get_dependency(name)
        for item in dependencies:
            if item not in openned:
                dependencies.remove(item); 
        return dependencies
    
class Parser:
    def __init__(self, sheet):
        self.sheet = sheet
        self.ctx={
            'name': None,
            'id': None,
            'subject_name_col': None,
            'credit_hours_col': None,
            'points_col': None,
            'score_col': None,
            'grade_col': None,
            'level': 0,
            'semester': 0,
        }
        self.m_row = 0
        self.m_col = 0

    def toggle_searching(self):
        self.searchinnng = not self.searchinnng

    def parse(self) -> json:

        max_row = self.sheet.max_row
        max_col = self.sheet.max_column
        subjects = []
        self.parse_id()
        self.parse_name()
        self.fill_ctx()
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = self.sheet.cell(row=row, column=col).value
                if cell_value == 'م':
                    row = row + 1
                    while self.sheet.cell(row=row, column=col).value:
                        if (self.sheet.cell(row=row, column=self.ctx.get('grade_col')).value == 'اعتذار'):
                            row = row + 1
                            continue
                        subjects.append(self.read_subject(row))
                        row = row + 1

                elif isinstance(cell_value, str) :
                    if re.search(r"المستوى/الفصل", cell_value):
                        pattern = r"المستوى\s+([^/\s]+).*?/الفصل(?:\s+الدراسي)?\s+(.+)$"
                        match = re.search(pattern, cell_value)
                        if match:
                            self.ctx['level'] = self.get_level(match.group(1))
                            self.ctx['semester'] = self.get_semester(match.group(2))

        return self.ctx.get('name'), pd.DataFrame(subjects)


    def get_level(self, cell_value: str) -> int:
        if cell_value == 'الأول' or cell_value == 'الاول':
            return 1
        elif cell_value == 'الثاني' or cell_value == 'الثانى':
            return 2    
        elif cell_value == 'الثالث':
            return 3
        elif cell_value == 'الرابع':
            return 4

    def get_semester(self, cell_value: str) -> int:
        if cell_value == 'الأول' or cell_value == 'الاول':
            return 1
        elif cell_value == 'الثاني' or cell_value == 'الثانى':
            return 2
        else:
            return 3

    def read_subject(self, row_index: int):
        return  {
            
            'name': self.sheet.cell(row=row_index, column=self.ctx['subject_name_col']).value,
            'credit_hours': float(self.sheet.cell(row=row_index, column=self.ctx['credit_hours_col']).value),
            'points': float(self.sheet.cell(row=row_index, column=self.ctx['points_col']).value) if isinstance(self.sheet.cell(row=row_index, column=self.ctx['points_col']).value, (int, float)) else 0.0,
            'score': float(self.sheet.cell(row=row_index, column=self.ctx['score_col']).value) if isinstance(self.sheet.cell(row=row_index, column=self.ctx['score_col']).value, (int, float)) else 0.0,
            'level': self.ctx['level'],
            'semester': self.ctx['semester'],
        }

    def parse_id(self,):
        for row in range(1, self.sheet.max_row + 1):
            for col in range(1, self.sheet.max_column + 1):
                cell_value = self.sheet.cell(row=row, column=col).value
                if isinstance(cell_value, str):
                    match = re.search(r"كود الطالب\s*:\s*(\d+)", cell_value)
                    if match:
                        self.ctx['id'] = match.group(1)
                        break
    def parse_name(self,):
        for row in range(1, self.sheet.max_row + 1):
            for col in range(1, self.sheet.max_column + 1):
                cell_value = self.sheet.cell(row=row, column=col).value
                if isinstance(cell_value, str):
                    match = re.search(r"أسم الطالب\s*:\s*(.+)", cell_value)
                    if match:
                        self.ctx['name'] = match.group(1)
                        break
    
    def fill_ctx(self,):
        for row in range(1, self.sheet.max_row + 1):
            for col in range(1, self.sheet.max_column + 1):
                cell_value = self.sheet.cell(row=row, column=col).value
                if isinstance(cell_value, str):
                    if cell_value == 'م':
                        self.ctx['m_col'] = col
                    elif cell_value == 'اسم المقرر':
                        self.ctx['subject_name_col'] = col
                    elif cell_value == 'ساعات':
                        self.ctx['credit_hours_col'] = col
                    elif cell_value == 'نقاط':
                        self.ctx['points_col'] = col
                    elif cell_value == 'درجة':
                        self.ctx['score_col'] = col
                    elif cell_value == 'تقدير':
                        self.ctx['grade_col'] = col


def parse_wb(wb_path:str ):
    workbook = xl.load_workbook(wb_path, data_only=True)
    rows = []
    

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        parser = Parser(sheet)
        student_name, df = parser.parse()
        #df = handle_zero_credit(df)
        df = handle_duplicates(df, passeed_only=True)
        #df = handle_zero_on_fail(df)
           
        openned = list(total_subjects - set(df.sort_values(by=["level", "semester"],ascending=[False, False])['name']))
       
        rank_map = dict()
        for name in openned:
            rank_map[name] = 0
        
        for name in next_year_lockers:
            if rank_map.get(name) != None:
                rank_map[name] = rank_map[name] + 1

 

        ndf= pd.DataFrame(rank_map.items(), columns=['subject_name', '#closings'])
        ndf['dependent subjects'] = ndf['subject_name'].apply(lambda subject: get_dependent_subjects(subject,openned))
        
        if(ndf['subject_name'].count() >= 1):   
            ndf.sort_values('#closings', ascending=False, ignore_index=True).to_excel(f'intermediate/{student_name}.xlsx')
